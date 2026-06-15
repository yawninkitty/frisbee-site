# competitions/views.py

import json
import re
import uuid
from datetime import datetime, timedelta

from django.contrib.auth.decorators import login_required
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.core.mail import send_mail
from django.http import HttpResponse, JsonResponse, Http404
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Q

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from urllib.parse import quote

from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

import os
from django.conf import settings

from users.models import Notification, User
from users.notifications import NotificationService

from .models import (
    Application, ApplicationEntry, Competition, CompetitionFile, Dog, Entry,
    EntryJudge, OrganizerContact, PaymentMethod, Result
)

from journal.models import Article


# ============================================================
# ПУБЛИЧНЫЕ СТРАНИЦЫ
# ============================================================

def index(request):
    """Главная — без крошек"""
    upcoming_competitions = Competition.objects.filter(
        status=Competition.STATUS_PUBLISHED,
        date__gte=timezone.now().date()
    ).order_by('date')[:8]

    recent_articles = Article.objects.filter(
        is_published=True
    ).order_by('-created_at')[:6]

    return render(request, 'competitions/index.html', {
        'upcoming_competitions': upcoming_competitions,
        'recent_articles': recent_articles,
    })


def competition_list(request):
    """Список соревнований"""
    competitions = Competition.objects.filter(
        status=Competition.STATUS_PUBLISHED
    ).order_by('date')
    return render(request, 'competitions/competition_list.html', {
        'competitions': competitions,
        'breadcrumbs': [
            {'name': 'Главная', 'url': reverse('competitions:index')},
            {'name': 'Соревнования', 'url': None},
        ],
    })


def competition_detail(request, pk):
    """Детальная страница соревнования"""
    competition = get_object_or_404(Competition, pk=pk)

    if competition.status == Competition.STATUS_DRAFT:
        if not request.user.is_authenticated or request.user != competition.organizer:
            raise Http404("Соревнование не найдено")

    upcoming_competitions = Competition.objects.filter(
        status=Competition.STATUS_PUBLISHED,
        date__gte=timezone.now().date()
    ).exclude(id=pk).order_by('date')[:8]

    recent_articles = Article.objects.filter(
        is_published=True
    ).order_by('-created_at')[:6]

    return render(request, 'competitions/competition_detail.html', {
        'competition': competition,
        'upcoming_competitions': upcoming_competitions,
        'recent_articles': recent_articles,
        'breadcrumbs': [
            {'name': 'Главная', 'url': reverse('competitions:index')},
            {'name': 'Соревнования', 'url': reverse('competitions:competition_list')},
            {'name': competition.title, 'url': None},
        ],
    })


def competition_results(request, pk):
    """Публичная страница результатов соревнования"""
    competition = get_object_or_404(
        Competition,
        pk=pk,
        status=Competition.STATUS_PUBLISHED
    )
    return render(request, 'competitions/competition_results.html', {
        'competition': competition,
        'entries': competition.entries.all(),
        'breadcrumbs': [
            {'name': 'Главная', 'url': reverse('competitions:index')},
            {'name': 'Соревнования', 'url': reverse('competitions:competition_list')},
            {'name': competition.title, 'url': reverse('competitions:competition_detail', args=[pk])},
            {'name': 'Результаты', 'url': None},
        ],
    })


def application_success(request, pk):
    """Страница успешной подачи заявки"""
    competition = get_object_or_404(Competition, pk=pk)

    recent_articles = Article.objects.filter(
        is_published=True
    ).order_by('-created_at')[:6]

    return render(request, 'competitions/application_success.html', {
        'competition': competition,
        'recent_articles': recent_articles,
        'breadcrumbs': [
            {'name': 'Главная', 'url': reverse('competitions:index')},
            {'name': 'Соревнования', 'url': reverse('competitions:competition_list')},
            {'name': competition.title, 'url': reverse('competitions:competition_detail', args=[pk])},
            {'name': 'Заявка отправлена', 'url': None},
        ],
    })


# ============================================================
# РАБОТА С ЗАЯВКАМИ
# ============================================================

@login_required
def register_for_competition(request, pk):
    """Подача заявки на соревнование"""
    competition = get_object_or_404(
        Competition,
        pk=pk,
        status=Competition.STATUS_PUBLISHED
    )

    if request.method == 'POST':
        if request.content_type == 'application/json':
            data = json.loads(request.body)
            applications_data = data.get('applications', [])
            created_applications = []

            for app_data in applications_data:
                dog_id = app_data.get('dog_id')
                selected_entries = app_data.get('entries', [])
                comment = app_data.get('comment', '')

                if not (dog_id and selected_entries):
                    continue

                dog = Dog.objects.get(id=dog_id)
                is_own_dog = dog.owner == request.user
                status = Application.STATUS_PENDING if is_own_dog else Application.STATUS_WAITING_OWNER

                application = Application.objects.create(
                    user=request.user,
                    dog_id=dog_id,
                    competition=competition,
                    comment=comment,
                    status=status
                )

                for entry_data in selected_entries:
                    entry_id = entry_data.get('entry_id')
                    is_out_of_class = entry_data.get('is_out_of_class', False)

                    if entry_id:
                        ApplicationEntry.objects.create(
                            application=application,
                            entry_id=entry_id,
                            is_out_of_class=is_out_of_class
                        )

                created_applications.append(application)

                NotificationService.send(
                    user=request.user,
                    type_=Notification.Type.APPLICATION_NEW,
                    title="Заявка отправлена",
                    message=f"Заявка на соревнование '{competition.title}' для собаки {dog.name} успешно отправлена.",
                    link=f"/competitions/{competition.id}/"
                )

                NotificationService.send(
                    user=competition.organizer,
                    type_=Notification.Type.APPLICATION_NEW,
                    title="Новая заявка",
                    message=f"Новая заявка от {request.user.get_full_name() or request.user.username} на соревнование '{competition.title}' для собаки {dog.name}.",
                    link=f"/competitions/organizer/competition/{competition.id}/manage/"
                )

                if not is_own_dog:
                    NotificationService.send(
                        user=dog.owner,
                        type_=Notification.Type.APPLICATION_NEW,
                        title="Заявка на вашу собаку",
                        message=f"Пользователь {request.user.get_full_name() or request.user.username} подал заявку с вашей собакой {dog.name} на соревнование '{competition.title}'. Требуется подтверждение.",
                        link=f"/users/profile/#applications"
                    )

            return render(request, 'competitions/application_success.html', {
                'competition': competition,
                'applications': created_applications,
                'recent_articles': Article.objects.filter(is_published=True).order_by('-created_at')[:6],
                'breadcrumbs': [
                    {'name': 'Главная', 'url': reverse('competitions:index')},
                    {'name': 'Соревнования', 'url': reverse('competitions:competition_list')},
                    {'name': competition.title, 'url': reverse('competitions:competition_detail', args=[pk])},
                    {'name': 'Заявка отправлена', 'url': None},
                ],
            })

    dogs = Dog.objects.filter(owner=request.user)
    entries = competition.entries.all()

    entries_by_discipline = {}
    for entry in entries:
        discipline = entry.discipline
        if discipline not in entries_by_discipline:
            entries_by_discipline[discipline] = []
        entries_by_discipline[discipline].append({
            'id': entry.id,
            'get_discipline_display': entry.get_discipline_display(),
            'get_sport_class_display': entry.get_sport_class_display(),
            'fee': str(entry.fee),
            'can_enter_without_class': entry.can_enter_without_class,
        })

    entries_by_discipline_json = json.dumps(entries_by_discipline, default=str)

    return render(request, 'competitions/register.html', {
        'competition': competition,
        'dogs': dogs,
        'entries_by_discipline': entries_by_discipline_json,
        'breadcrumbs': [
            {'name': 'Главная', 'url': reverse('competitions:index')},
            {'name': 'Соревнования', 'url': reverse('competitions:competition_list')},
            {'name': competition.title, 'url': reverse('competitions:competition_detail', args=[pk])},
            {'name': 'Подать заявку', 'url': None},
        ],
    })


# ============================================================
# УПРАВЛЕНИЕ ЗАЯВКАМИ (ОРГАНИЗАТОР)
# ============================================================

@login_required
def update_application_status(request):
    """API — без крошек"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Метод не поддерживается'})

    data = json.loads(request.body)
    app_id = data.get('app_id')
    status = data.get('status')
    organizer_comment = data.get('organizer_comment', '')

    try:
        application = Application.objects.get(id=app_id, competition__organizer=request.user)

        if application.status == Application.STATUS_CANCELLED:
            return JsonResponse({'success': False, 'error': 'Нельзя изменить статус отменённой заявки'})

        if Result.objects.filter(application_entry__application=application).exists():
            return JsonResponse({'success': False, 'error': 'Нельзя изменить статус заявки, по которой уже есть результаты'})

        old_status = application.status
        application.status = status
        application.organizer_comment = organizer_comment
        application.save()

        status_messages = {
            Application.STATUS_APPROVED: "одобрена",
            Application.STATUS_REJECTED: "отклонена",
            Application.STATUS_CANCELLED: "отменена",
        }
        if status in status_messages and old_status != status:
            status_text = status_messages[status]
            message = f"Ваша заявка на соревнование '{application.competition.title}' для собаки {application.dog.name} {status_text}."
            if status == Application.STATUS_REJECTED and organizer_comment:
                message += f"\nПричина: {organizer_comment}"

            NotificationService.send(
                user=application.user,
                type_=Notification.Type.APPLICATION_STATUS,
                title=f"Заявка {status_text}",
                message=message,
                link=f"/users/profile/#applications"
            )

        if old_status != 'approved' and status == 'approved':
            send_mail(
                f'Заявка на {application.competition.title} одобрена',
                f'Ваша заявка на соревнование "{application.competition.title}" одобрена!',
                'noreply@example.com',
                [application.user.email],
                fail_silently=True,
            )
        elif old_status != 'rejected' and status == 'rejected' and organizer_comment:
            send_mail(
                f'Заявка на {application.competition.title} отклонена',
                f'Ваша заявка на соревнование "{application.competition.title}" отклонена.\n\nПричина: {organizer_comment}',
                'noreply@example.com',
                [application.user.email],
                fail_silently=True,
            )

        return JsonResponse({'success': True})
    except Application.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Заявка не найдена'})


@login_required
def update_application_payment_status(request):
    """API — без крошек"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Метод не поддерживается'})

    data = json.loads(request.body)
    app_id = data.get('app_id')
    payment_status = data.get('payment_status')

    try:
        application = Application.objects.get(id=app_id, competition__organizer=request.user)

        if application.status == Application.STATUS_CANCELLED:
            return JsonResponse({'success': False, 'error': 'Нельзя изменить статус отменённой заявки'})

        if Result.objects.filter(application_entry__application=application).exists():
            return JsonResponse({'success': False, 'error': 'Нельзя изменить статус заявки, по которой уже есть результаты'})

        old_payment_status = application.payment_status
        application.payment_status = payment_status
        application.save()

        if old_payment_status != payment_status:
            payment_messages = {
                Application.PAYMENT_STATUS_PAID: "подтверждена",
                Application.PAYMENT_STATUS_PENDING: "на проверке",
            }
            if payment_status in payment_messages:
                status_text = payment_messages[payment_status]
                NotificationService.send(
                    user=application.user,
                    type_=Notification.Type.APPLICATION_PAYMENT,
                    title=f"Оплата {status_text}",
                    message=f"Статус оплаты заявки на соревнование '{application.competition.title}' для собаки {application.dog.name} изменён на: {application.get_payment_status_display()}.",
                    link=f"/users/profile/#applications"
                )

        return JsonResponse({'success': True})
    except Application.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Заявка не найдена'})


# ============================================================
# РАБОТА С РЕЗУЛЬТАТАМИ
# ============================================================

def recalc_places(entry):
    Result.objects.filter(application_entry__entry=entry).update(place=None)

    active_results = Result.objects.filter(
        application_entry__entry=entry,
        status=Result.STATUS_ACTIVE
    ).order_by('-sort_value')

    for place, result in enumerate(active_results, start=1):
        result.place = place
        result.save()


def ensure_results_for_entry(entry):
    for app_entry in ApplicationEntry.objects.filter(entry=entry):
        if (app_entry.application.status == Application.STATUS_APPROVED and
                app_entry.application.payment_status == Application.PAYMENT_STATUS_PAID):
            Result.objects.get_or_create(
                application_entry=app_entry,
                defaults={'data': {}, 'sort_value': 0, 'place': None, 'status': Result.STATUS_ACTIVE}
            )


@login_required
def update_result_data(request):
    """API — без крошек"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Метод не поддерживается'})

    data = json.loads(request.body)
    result_id = data.get('result_id')
    new_data = data.get('data')
    new_status = data.get('status')

    try:
        result = Result.objects.get(id=result_id)
        competition = result.application_entry.entry.competition
        entry = result.application_entry.entry

        is_organizer = competition.organizer == request.user
        is_assigned_judge = entry.assigned_judges.filter(user=request.user).exists()

        if not (is_organizer or is_assigned_judge):
            return JsonResponse({'success': False, 'error': 'Нет прав'})

        old_data = result.data
        result.data = new_data

        if new_status:
            result.status = new_status

        discipline = entry.discipline
        if discipline == 'bullseye':
            result.sort_value = new_data.get('total', 0)
        elif discipline == 'distance':
            result.sort_value = new_data.get('best', 0)
        elif discipline == 'accuracy':
            result.sort_value = new_data.get('top_five_sum', 0)

        result.save()
        recalc_places(entry)

        if old_data != new_data:
            application = result.application_entry.application
            NotificationService.send(
                user=application.user,
                type_=Notification.Type.RESULT_READY,
                title="Результат опубликован",
                message=f"Опубликован результат для собаки {application.dog.name} в дисциплине {entry.get_discipline_display()} ({entry.get_sport_class_display()}) на соревновании '{competition.title}'.",
                link=f"/competitions/{competition.id}/results/"
            )

        return JsonResponse({'success': True})
    except Result.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Результат не найден'})


@login_required
def update_result_status(request):
    """API — без крошек"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Метод не поддерживается'})

    data = json.loads(request.body)
    result_id = data.get('result_id')
    new_status = data.get('status')

    try:
        result = Result.objects.get(id=result_id)
        competition = result.application_entry.entry.competition
        entry = result.application_entry.entry

        is_organizer = competition.organizer == request.user
        is_assigned_judge = entry.assigned_judges.filter(user=request.user).exists()

        if not (is_organizer or is_assigned_judge):
            return JsonResponse({'success': False, 'error': 'Нет прав'})

        result.status = new_status
        result.save()
        recalc_places(entry)

        if new_status == Result.STATUS_DISQUALIFIED:
            application = result.application_entry.application
            NotificationService.send(
                user=application.user,
                type_=Notification.Type.RESULT_READY,
                title="Изменение статуса участника",
                message=f"Ваш результат для собаки {application.dog.name} на соревновании '{competition.title}' изменён на: дисквалифицирован.",
                link=f"/competitions/{competition.id}/results/"
            )

        return JsonResponse({'success': True})
    except Result.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Результат не найден'})


@login_required
def enter_results(request, pk, entry_id):
    """Страница ввода результатов"""
    competition = get_object_or_404(Competition, pk=pk)
    entry = get_object_or_404(Entry, pk=entry_id, competition=competition)

    is_organizer = competition.organizer == request.user
    is_assigned_judge = entry.assigned_judges.filter(user=request.user).exists()

    if not (is_organizer or is_assigned_judge):
        return redirect('competitions:competition_detail', pk=pk)

    results = Result.objects.filter(
        application_entry__entry=entry,
        application_entry__application__status=Application.STATUS_APPROVED,
        application_entry__application__payment_status=Application.PAYMENT_STATUS_PAID
    ).select_related('application_entry__application__user', 'application_entry__application__dog')

    participants = []
    for result in results:
        app = result.application_entry.application
        participants.append({
            'result_id': result.id,
            'user_name': app.user.get_full_name() or app.user.username,
            'dog_name': app.dog.name,
            'data': result.data,
            'status': result.status,
            'sort_value': result.sort_value
        })

    return render(request, 'competitions/enter_results.html', {
        'competition': competition,
        'entry': entry,
        'participants': participants,
        'discipline': entry.discipline,
        'is_organizer': is_organizer,
        'is_judge': is_assigned_judge,
        'breadcrumbs': [
            {'name': 'Главная', 'url': reverse('competitions:index')},
            {'name': 'Кабинет организатора', 'url': reverse('competitions:organizer_dashboard')},
            {'name': competition.title, 'url': reverse('competitions:manage_competition', args=[pk])},
            {'name': f'{entry.get_discipline_display()} — {entry.get_sport_class_display()}', 'url': None},
        ],
    })


# ============================================================
# УПРАВЛЕНИЕ СОРЕВНОВАНИЯМИ (ОРГАНИЗАТОР)
# ============================================================

@login_required
def organizer_dashboard(request):
    """Дашборд организатора/судьи"""
    organized_comps = Competition.objects.filter(organizer=request.user).order_by('-date')
    judging_comps = Competition.objects.filter(
        entries__assigned_judges__user=request.user
    ).distinct().order_by('-date')

    for comp in organized_comps:
        comp.total_applications = Application.objects.filter(competition=comp).count()
        comp.paid_applications = Application.objects.filter(
            competition=comp, payment_status=Application.PAYMENT_STATUS_PAID
        ).count()
        comp.user_role = 'organizer'
        comp.can_edit = True
        comp.can_manage_results = True

    for comp in judging_comps:
        comp.user_role = 'judge'
        comp.can_edit = False
        comp.can_manage_results = True
        comp.judging_entries = comp.entries.filter(assigned_judges__user=request.user)

    all_comps = {comp.id: comp for comp in organized_comps}
    for comp in judging_comps:
        if comp.id not in all_comps:
            all_comps[comp.id] = comp

    competitions = list(all_comps.values())

    return render(request, 'competitions/organizer_dashboard.html', {
        'organized_competitions': [c for c in competitions if c.user_role == 'organizer'],
        'judging_competitions': [c for c in competitions if c.user_role == 'judge'],
        'breadcrumbs': [
            {'name': 'Главная', 'url': reverse('competitions:index')},
            {'name': 'Кабинет организатора', 'url': None},
        ],
    })


@login_required
def manage_competition(request, pk):
    """Управление конкретным соревнованием"""
    competition = get_object_or_404(Competition, pk=pk)

    is_main_organizer = competition.organizer == request.user
    is_judge = Entry.objects.filter(
        competition=competition, assigned_judges__user=request.user
    ).exists()

    if not (is_main_organizer or is_judge):
        return redirect('competitions:organizer_dashboard')

    applications = []
    if is_main_organizer:
        applications = Application.objects.filter(
            competition=competition
        ).exclude(
            status=Application.STATUS_WAITING_OWNER
        ).select_related('user', 'dog').prefetch_related(
            'application_entries__entry'
        ).order_by('-created_at')

        for app in applications:
            app.entries_list = app.application_entries.select_related('entry').all()
            app.has_results = Result.objects.filter(application_entry__application=app).exists()

    entries_for_display = competition.entries.all() if is_main_organizer else competition.entries.filter(
        assigned_judges__user=request.user
    )

    # Сортируем в порядке: Буллсай → Дальность → Точность, Новички → Прогресс → Открытый
    discipline_order = ['bullseye', 'distance', 'accuracy']
    class_order = ['novice', 'progress', 'open']

    entries_for_display = sorted(
        entries_for_display,
        key=lambda e: (
            discipline_order.index(e.discipline) if e.discipline in discipline_order else 99,
            class_order.index(e.sport_class) if e.sport_class in class_order else 99
        )
    )

    return render(request, 'competitions/manage_competition.html', {
        'competition': competition,
        'applications': applications,
        'entries_for_display': entries_for_display,
        'is_main_organizer': is_main_organizer,
        'is_judge': is_judge,
        'breadcrumbs': [
            {'name': 'Главная', 'url': reverse('competitions:index')},
            {'name': 'Кабинет организатора', 'url': reverse('competitions:organizer_dashboard')},
            {'name': competition.title, 'url': None},
        ],
    })


@login_required
def create_competition(request):
    """Создание нового соревнования"""
    if not request.user.profile.is_organizer:
        return redirect('competitions:index')

    payment_methods = PaymentMethod.objects.all()
    organizers = User.objects.filter(profile__is_organizer=True).exclude(id=request.user.id)

    if request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        title = request.POST.get('title')
        date = request.POST.get('date')
        registration_deadline = request.POST.get('registration_deadline')
        location = request.POST.get('location')
        location_type = request.POST.get('location_type')
        city = request.POST.get('city', '')
        description = request.POST.get('description', '')
        payment_deadline = request.POST.get('payment_deadline') or None
        time = request.POST.get('time') or None
        status = request.POST.get('status', Competition.STATUS_DRAFT)

        competition = Competition.objects.create(
            title=title or "Новое соревнование",
            date=date if date else None,
            registration_deadline=registration_deadline if registration_deadline else None,
            location=location or "",
            location_type=location_type or "outdoor",
            city=city or "",
            description=description or "",
            organizer=request.user,
            payment_deadline=payment_deadline if payment_deadline else None,
            time=time if time else None,
            status=status
        )

        if status == Competition.STATUS_DRAFT:
            competition.payment_methods.set(request.POST.getlist('payment_methods'))
            for ct, cv in zip(request.POST.getlist('contact_type[]'), request.POST.getlist('contact_value[]')):
                if ct and cv:
                    OrganizerContact.objects.create(competition=competition, contact_type=ct, value=cv)
            for temp_path in request.POST.getlist('new_files[]'):
                if default_storage.exists(temp_path):
                    content = default_storage.open(temp_path).read()
                    raw_name = temp_path.split('/')[-1]
                    original_name = raw_name.split('___', 1)[1] if '___' in raw_name else raw_name
                    file_name = original_name
                    counter = 1
                    while default_storage.exists(f'competition_files/{competition.id}/{file_name}'):
                        name_part, ext_part = os.path.splitext(original_name)
                        file_name = f'{name_part}_{counter}{ext_part}'
                        counter += 1
                    new_path = f'competition_files/{competition.id}/{file_name}'
                    default_storage.save(new_path, ContentFile(content))
                    CompetitionFile.objects.create(competition=competition, title=file_name, file=new_path)
                    default_storage.delete(temp_path)
            entries_data = {}
            for key, value in request.POST.items():
                if key.startswith('entries['):
                    match = re.match(r'entries\[(\d+)\]\[(\w+)\]', key)
                    if match:
                        idx, field = int(match.group(1)), match.group(2)
                        entries_data.setdefault(idx, {})[field] = value
            for entry_data in entries_data.values():
                discipline = entry_data.get('discipline', '')
                sport_class = entry_data.get('sport_class', '')
                judge_id = entry_data.get('judge_id', '')
                judge_name = entry_data.get('judge_name', '')
                fee = entry_data.get('fee', 0)
                can_enter_without_class = entry_data.get('can_enter_without_class', 'false').lower() == 'true'
                entry = Entry.objects.create(competition=competition, discipline=discipline, sport_class=sport_class, judge_name=judge_name, fee=fee, can_enter_without_class=can_enter_without_class)
                if judge_id:
                    try:
                        EntryJudge.objects.create(entry=entry, user=User.objects.get(id=judge_id))
                    except User.DoesNotExist:
                        pass
            return JsonResponse({'success': True, 'competition_id': competition.id})

        if not all([title, date, registration_deadline, location, location_type]):
            competition.delete()
            return JsonResponse({'success': False, 'error': 'Заполните все обязательные поля'})

        contact_types = request.POST.getlist('contact_type[]')
        contact_values = request.POST.getlist('contact_value[]')
        has_contact = any(ct and cv for ct, cv in zip(contact_types, contact_values))
        if not has_contact:
            competition.delete()
            return JsonResponse({'success': False, 'error': 'Добавьте хотя бы один контакт для связи'})

        competition.payment_methods.set(request.POST.getlist('payment_methods'))
        for ct, cv in zip(contact_types, contact_values):
            if ct and cv:
                OrganizerContact.objects.create(competition=competition, contact_type=ct, value=cv)
        for temp_path in request.POST.getlist('new_files[]'):
            if default_storage.exists(temp_path):
                content = default_storage.open(temp_path).read()
                raw_name = temp_path.split('/')[-1]
                original_name = raw_name.split('___', 1)[1] if '___' in raw_name else raw_name
                file_name = original_name
                counter = 1
                while default_storage.exists(f'competition_files/{competition.id}/{file_name}'):
                    name_part, ext_part = os.path.splitext(original_name)
                    file_name = f'{name_part}_{counter}{ext_part}'
                    counter += 1
                new_path = f'competition_files/{competition.id}/{file_name}'
                default_storage.save(new_path, ContentFile(content))
                CompetitionFile.objects.create(competition=competition, title=file_name, file=new_path)
                default_storage.delete(temp_path)
        entries_data = {}
        for key, value in request.POST.items():
            if key.startswith('entries['):
                match = re.match(r'entries\[(\d+)\]\[(\w+)\]', key)
                if match:
                    idx, field = int(match.group(1)), match.group(2)
                    entries_data.setdefault(idx, {})[field] = value
        for entry_data in entries_data.values():
            discipline = entry_data.get('discipline')
            sport_class = entry_data.get('sport_class')
            judge_id = entry_data.get('judge_id')
            judge_name = entry_data.get('judge_name', '')
            fee = entry_data.get('fee', 0)
            can_enter_without_class = entry_data.get('can_enter_without_class', 'false').lower() == 'true'
            if not all([discipline, sport_class]) or not (judge_id or judge_name):
                continue
            entry = Entry.objects.create(competition=competition, discipline=discipline, sport_class=sport_class, judge_name=judge_name, fee=fee, can_enter_without_class=can_enter_without_class)
            if judge_id and judge_id != '':
                try:
                    EntryJudge.objects.create(entry=entry, user=User.objects.get(id=judge_id))
                except User.DoesNotExist:
                    pass

        return JsonResponse({'success': True, 'competition_id': competition.id})

    return render(request, 'competitions/competition_form.html', {
        'payment_methods': payment_methods,
        'organizers': organizers,
        'breadcrumbs': [
            {'name': 'Главная', 'url': reverse('competitions:index')},
            {'name': 'Кабинет организатора', 'url': reverse('competitions:organizer_dashboard')},
            {'name': 'Создать соревнование', 'url': None},
        ],
    })


@login_required
def edit_competition(request, pk):
    """Редактирование соревнования"""
    competition = get_object_or_404(Competition, pk=pk, organizer=request.user)

    if not request.user.profile.is_organizer or competition.organizer != request.user:
        return redirect('competitions:organizer_dashboard')

    payment_methods = PaymentMethod.objects.all()
    organizers = User.objects.filter(profile__is_organizer=True).exclude(id=request.user.id)

    if request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        try:
            if request.POST.get('save_as_draft') == 'true':
                if request.POST.get('title'): competition.title = request.POST.get('title')
                if request.POST.get('date'): competition.date = request.POST.get('date')
                if request.POST.get('registration_deadline'): competition.registration_deadline = request.POST.get('registration_deadline')
                if request.POST.get('location'): competition.location = request.POST.get('location')
                if request.POST.get('location_type'): competition.location_type = request.POST.get('location_type')
                if request.POST.get('city'): competition.city = request.POST.get('city')
                if request.POST.get('description') is not None: competition.description = request.POST.get('description')
                if request.POST.get('payment_deadline'): competition.payment_deadline = request.POST.get('payment_deadline')
                if request.POST.get('time'): competition.time = request.POST.get('time')
                competition.status = Competition.STATUS_DRAFT
                competition.save()
                if request.POST.getlist('payment_methods'): competition.payment_methods.set(request.POST.getlist('payment_methods'))
                competition.contacts.all().delete()
                for ct, cv in zip(request.POST.getlist('contact_type[]'), request.POST.getlist('contact_value[]')):
                    if ct and cv: OrganizerContact.objects.create(competition=competition, contact_type=ct, value=cv)
                for file_id in request.POST.getlist('delete_files[]'): CompetitionFile.objects.filter(id=file_id, competition=competition).delete()
                for temp_path in request.POST.getlist('new_files[]'):
                    if default_storage.exists(temp_path):
                        content = default_storage.open(temp_path).read()
                        raw_name = temp_path.split('/')[-1]
                        original_name = raw_name.split('___', 1)[1] if '___' in raw_name else raw_name
                        file_name = original_name
                        counter = 1
                        while default_storage.exists(f'competition_files/{competition.id}/{file_name}'):
                            name_part, ext_part = os.path.splitext(original_name)
                            file_name = f'{name_part}_{counter}{ext_part}'
                            counter += 1
                        new_path = f'competition_files/{competition.id}/{file_name}'
                        default_storage.save(new_path, ContentFile(content))
                        CompetitionFile.objects.create(competition=competition, title=file_name, file=new_path)
                        default_storage.delete(temp_path)
                for entry_id in request.POST.getlist('delete_entries[]'): Entry.objects.filter(id=entry_id, competition=competition).delete()
                entries_data = {}
                for key, value in request.POST.items():
                    if key.startswith('entries['):
                        match = re.match(r'entries\[(\d+)\]\[(\w+)\]', key)
                        if match:
                            idx, field = int(match.group(1)), match.group(2)
                            entries_data.setdefault(idx, {})[field] = value
                for entry_data in entries_data.values():
                    discipline = entry_data.get('discipline', '')
                    sport_class = entry_data.get('sport_class', '')
                    judge_id = entry_data.get('judge_id', '')
                    judge_name = entry_data.get('judge_name', '')
                    fee = entry_data.get('fee', 0)
                    can_enter_without_class = entry_data.get('can_enter_without_class', 'false').lower() == 'true'
                    if 'id' in entry_data and entry_data['id']:
                        try:
                            entry = Entry.objects.get(id=entry_data['id'], competition=competition)
                            if discipline: entry.discipline = discipline
                            if sport_class: entry.sport_class = sport_class
                            entry.judge_name = judge_name
                            entry.fee = fee
                            entry.can_enter_without_class = can_enter_without_class
                            entry.save()
                            EntryJudge.objects.filter(entry=entry).delete()
                            if judge_id:
                                try: EntryJudge.objects.create(entry=entry, user=User.objects.get(id=judge_id))
                                except User.DoesNotExist: pass
                        except Entry.DoesNotExist: pass
                    else:
                        existing = Entry.objects.filter(competition=competition, discipline=discipline, sport_class=sport_class).first()
                        if existing:
                            existing.judge_name = judge_name
                            existing.fee = fee
                            existing.can_enter_without_class = can_enter_without_class
                            existing.save()
                            EntryJudge.objects.filter(entry=existing).delete()
                            if judge_id:
                                try: EntryJudge.objects.create(entry=existing, user=User.objects.get(id=judge_id))
                                except User.DoesNotExist: pass
                        else:
                            entry = Entry.objects.create(competition=competition, discipline=discipline, sport_class=sport_class, judge_name=judge_name, fee=fee, can_enter_without_class=can_enter_without_class)
                            if judge_id:
                                try: EntryJudge.objects.create(entry=entry, user=User.objects.get(id=judge_id))
                                except User.DoesNotExist: pass
                return JsonResponse({'success': True})

            competition.title = request.POST.get('title')
            competition.date = request.POST.get('date')
            competition.registration_deadline = request.POST.get('registration_deadline')
            competition.location = request.POST.get('location')
            competition.location_type = request.POST.get('location_type')
            competition.city = request.POST.get('city', '')
            competition.description = request.POST.get('description', '')
            competition.payment_deadline = request.POST.get('payment_deadline') or None
            competition.time = request.POST.get('time') or None
            new_status = request.POST.get('status')
            if new_status in [Competition.STATUS_DRAFT, Competition.STATUS_PUBLISHED]: competition.status = new_status
            competition.save()
            competition.payment_methods.set(request.POST.getlist('payment_methods'))
            competition.contacts.all().delete()
            for ct, cv in zip(request.POST.getlist('contact_type[]'), request.POST.getlist('contact_value[]')):
                if ct and cv: OrganizerContact.objects.create(competition=competition, contact_type=ct, value=cv)
            for file_id in request.POST.getlist('delete_files[]'): CompetitionFile.objects.filter(id=file_id, competition=competition).delete()
            for temp_path in request.POST.getlist('new_files[]'):
                if default_storage.exists(temp_path):
                    content = default_storage.open(temp_path).read()
                    raw_name = temp_path.split('/')[-1]
                    original_name = raw_name.split('___', 1)[1] if '___' in raw_name else raw_name
                    file_name = original_name
                    counter = 1
                    while default_storage.exists(f'competition_files/{competition.id}/{file_name}'):
                        name_part, ext_part = os.path.splitext(original_name)
                        file_name = f'{name_part}_{counter}{ext_part}'
                        counter += 1
                    new_path = f'competition_files/{competition.id}/{file_name}'
                    default_storage.save(new_path, ContentFile(content))
                    CompetitionFile.objects.create(competition=competition, title=file_name, file=new_path)
                    default_storage.delete(temp_path)
            for entry_id in request.POST.getlist('delete_entries[]'): Entry.objects.filter(id=entry_id, competition=competition).delete()
            entries_data = {}
            for key, value in request.POST.items():
                if key.startswith('entries['):
                    match = re.match(r'entries\[(\d+)\]\[(\w+)\]', key)
                    if match:
                        idx, field = int(match.group(1)), match.group(2)
                        entries_data.setdefault(idx, {})[field] = value
            for entry_data in entries_data.values():
                judge_id = entry_data.get('judge_id', '')
                judge_name = entry_data.get('judge_name', '')
                can_enter_without_class = entry_data.get('can_enter_without_class', 'false').lower() == 'true'
                if 'id' in entry_data and entry_data['id']:
                    entry = Entry.objects.get(id=entry_data['id'], competition=competition)
                    entry.discipline = entry_data['discipline']
                    entry.sport_class = entry_data['sport_class']
                    entry.judge_name = judge_name
                    entry.fee = entry_data.get('fee', 0)
                    entry.can_enter_without_class = can_enter_without_class
                    entry.save()
                    EntryJudge.objects.filter(entry=entry).delete()
                else:
                    existing = Entry.objects.filter(competition=competition, discipline=entry_data['discipline'], sport_class=entry_data['sport_class']).first()
                    if existing:
                        entry = existing
                        entry.judge_name = judge_name
                        entry.fee = entry_data.get('fee', 0)
                        entry.can_enter_without_class = can_enter_without_class
                        entry.save()
                        EntryJudge.objects.filter(entry=entry).delete()
                    else:
                        entry = Entry.objects.create(competition=competition, discipline=entry_data['discipline'], sport_class=entry_data['sport_class'], judge_name=judge_name, fee=entry_data.get('fee', 0), can_enter_without_class=can_enter_without_class)
                if judge_id and judge_id != '':
                    try:
                        judge_user = User.objects.get(id=judge_id)
                        EntryJudge.objects.create(entry=entry, user=judge_user)
                        if judge_user != request.user:
                            NotificationService.send(
                                user=judge_user, type_=Notification.Type.SYSTEM,
                                title="Назначение судьёй",
                                message=f"Вас назначили судьёй на соревнование '{competition.title}' в дисциплине {entry.get_discipline_display()} ({entry.get_sport_class_display()}).",
                                link=f"/organizer/competition/{competition.id}/manage/?tab=results"
                            )
                    except User.DoesNotExist: pass
            return JsonResponse({'success': True})
        except Exception as e:
            import traceback; traceback.print_exc()
            return JsonResponse({'success': False, 'error': str(e)})

    entries_data = []
    for entry in competition.entries.all():
        first_judge = entry.assigned_judges.first()
        entries_data.append({
            'id': entry.id, 'discipline': entry.discipline, 'sport_class': entry.sport_class,
            'judge_id': first_judge.user.id if first_judge else None, 'judge_name': entry.judge_name,
            'fee': float(entry.fee), 'can_enter_without_class': entry.can_enter_without_class,
        })
    return render(request, 'competitions/competition_form.html', {
        'competition': competition, 'payment_methods': payment_methods, 'organizers': organizers,
        'selected_payment_methods': list(competition.payment_methods.values_list('id', flat=True)),
        'contacts': competition.contacts.all(), 'files': competition.files.all(),
        'entries_data': json.dumps(entries_data),
        'breadcrumbs': [
            {'name': 'Главная', 'url': reverse('competitions:index')},
            {'name': 'Кабинет организатора', 'url': reverse('competitions:organizer_dashboard')},
            {'name': competition.title, 'url': reverse('competitions:manage_competition', args=[pk])},
            {'name': 'Редактировать', 'url': None},
        ],
    })


@login_required
def delete_competition(request, pk):
    """API — без крошек"""
    competition = get_object_or_404(Competition, pk=pk, organizer=request.user)
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Метод не поддерживается'})
    active_applications = Application.objects.filter(competition=competition).exclude(status__in=[Application.STATUS_CANCELLED, Application.STATUS_REJECTED]).exists()
    if active_applications:
        return JsonResponse({'success': False, 'error': 'Нельзя удалить соревнование с активными заявками.'})
    competition.delete()
    return JsonResponse({'success': True})


@login_required
def publish_competition(request, pk):
    """API — без крошек"""
    competition = get_object_or_404(Competition, pk=pk, organizer=request.user)
    if request.method == 'POST':
        errors = []
        if not competition.title or competition.title == "Новое соревнование": errors.append('Название')
        if not competition.date: errors.append('Дата проведения')
        if not competition.registration_deadline: errors.append('Дата окончания регистрации')
        if not competition.location: errors.append('Место проведения')
        if not competition.contacts.exists(): errors.append('Хотя бы один контакт')
        if not competition.entries.exists(): errors.append('Хотя бы один зачёт')
        if errors:
            return JsonResponse({'success': False, 'error': f'Заполните обязательные поля: {", ".join(errors)}'})
        competition.status = Competition.STATUS_PUBLISHED
        competition.save()
        return JsonResponse({'success': True})
    return JsonResponse({'success': False, 'error': 'Метод не поддерживается'})


# ============================================================
# ЭКСПОРТ РЕЗУЛЬТАТОВ
# ============================================================

@login_required
def export_results(request, pk):
    """Экспорт результатов соревнования в PDF или Excel"""
    competition = get_object_or_404(Competition, pk=pk)

    is_organizer = competition.organizer == request.user
    is_judge = Entry.objects.filter(competition=competition, assigned_judges__user=request.user).exists()

    if not (is_organizer or is_judge):
        return JsonResponse({'success': False, 'error': 'Нет прав'}, status=403)

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Метод не поддерживается'}, status=405)

    try:
        data = json.loads(request.body)
        export_format = data.get('format', 'pdf')
        selected_entries = data.get('entries', [])

        if not selected_entries:
            return JsonResponse({'success': False, 'error': 'Не выбрано ни одного зачёта'})

        entries = Entry.objects.filter(id__in=selected_entries, competition=competition)

        if export_format == 'excel':
            return _export_to_excel(competition, entries)
        else:
            return _export_to_pdf(competition, entries)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Неверный формат данных'}, status=400)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


def _export_to_excel(competition, entries):
    """Экспорт в Excel (.xlsx)"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
        from urllib.parse import quote

        wb = Workbook()
        wb.remove(wb.active)

        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for entry in entries:
            sheet_name = f"{entry.get_discipline_display()} - {entry.get_sport_class_display()}"[:31]
            ws = wb.create_sheet(sheet_name)

            # Заголовок
            ws.merge_cells('A1:E1')
            title_cell = ws['A1']
            title_cell.value = f"Результаты соревнования: {competition.title}"
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = Alignment(horizontal="center")

            ws['A3'] = f"Дисциплина: {entry.get_discipline_display()}"
            ws['A4'] = f"Класс: {entry.get_sport_class_display()}"
            ws['A5'] = f"Дата: {competition.date}"

            results = Result.objects.filter(
                application_entry__entry=entry,
                application_entry__application__status=Application.STATUS_APPROVED,
                application_entry__application__payment_status=Application.PAYMENT_STATUS_PAID
            ).select_related(
                'application_entry__application__user',
                'application_entry__application__dog'
            ).order_by('place')

            # Заголовки таблицы
            headers = ['Место', 'Спортсмен', 'Собака', 'Статус']
            if entry.discipline == 'bullseye':
                headers.extend(['Броски', 'Сумма', 'Кол-во бросков'])
            elif entry.discipline == 'distance':
                headers.extend(['Попытка 1 (м)', 'Попытка 2 (м)', 'Попытка 3 (м)', 'Последний шанс (м)', 'Лучший (м)'])
            elif entry.discipline == 'accuracy':
                headers.extend(['Броски', 'Сумма 5 лучших', 'Общая сумма'])

            # Записываем заголовки
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=7, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border

            # Записываем данные
            row = 8
            for result in results:
                app = result.application_entry.application
                ws.cell(row=row, column=1, value=result.place or '—').border = border
                ws.cell(row=row, column=2, value=app.user.get_full_name() or app.user.username).border = border
                ws.cell(row=row, column=3, value=app.dog.name).border = border
                ws.cell(row=row, column=4, value='Активен' if result.status == 'active' else 'Снят').border = border

                col = 5
                if entry.discipline == 'bullseye':
                    throws = result.data.get('throws', [])
                    ws.cell(row=row, column=col, value=', '.join(map(str, throws)) if throws else '—').border = border
                    ws.cell(row=row, column=col + 1, value=result.data.get('total', '—')).border = border
                    ws.cell(row=row, column=col + 2, value=result.data.get('throw_count', '—')).border = border
                elif entry.discipline == 'distance':
                    attempts = result.data.get('attempts', [0, 0, 0])
                    ws.cell(row=row, column=col, value=attempts[0]).border = border
                    ws.cell(row=row, column=col + 1, value=attempts[1]).border = border
                    ws.cell(row=row, column=col + 2, value=attempts[2]).border = border
                    ws.cell(row=row, column=col + 3, value=result.data.get('last_chance', 0)).border = border
                    ws.cell(row=row, column=col + 4, value=result.data.get('best', 0)).border = border
                elif entry.discipline == 'accuracy':
                    throws = result.data.get('throws', [])
                    ws.cell(row=row, column=col, value=', '.join(map(str, throws)) if throws else '—').border = border
                    ws.cell(row=row, column=col + 1, value=result.data.get('top_five_sum', '—')).border = border
                    ws.cell(row=row, column=col + 2, value=result.data.get('total_sum', '—')).border = border
                row += 1

            # Авто-ширина колонок
            for col_idx in range(1, len(headers) + 1):
                max_length = 0
                col_letter = get_column_letter(col_idx)
                for row_idx in range(1, row):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value:
                        try:
                            max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[col_letter].width = adjusted_width

        # Формируем ответ с правильным content_type
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        # Очищаем название файла
        safe_title = "".join(c for c in competition.title if c.isalnum() or c in (' ', '-', '_')).rstrip()
        filename = f"results_{safe_title}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        # Ключевое исправление: используем filename* для UTF-8
        encoded_filename = quote(filename)
        response['Content-Disposition'] = f"attachment; filename=\"{filename}\"; filename*=UTF-8''{encoded_filename}"

        wb.save(response)
        return response

    except Exception as e:
        import traceback
        traceback.print_exc()
        return HttpResponse(f'Ошибка при создании Excel: {str(e)}', status=500)


def _export_to_pdf(competition, entries):
    """Экспорт в PDF через reportlab с таблицами"""

    response = HttpResponse(content_type='application/pdf')
    safe_title = "".join(c for c in competition.title if c.isalnum() or c in (' ', '-', '_')).rstrip()
    response[
        'Content-Disposition'] = f'attachment; filename="results_{safe_title}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'

    # Регистрируем шрифт
    font_path = os.path.join(settings.BASE_DIR, 'static', 'fonts', 'arial.ttf')
    if not os.path.exists(font_path):
        font_path = os.path.join(settings.BASE_DIR, 'static', 'fonts', 'DejaVuSans.ttf')

    try:
        pdfmetrics.registerFont(TTFont('RussianFont', font_path))
        font_name = 'RussianFont'
    except:
        font_name = 'Helvetica'

    # Создаём документ
    doc = SimpleDocTemplate(response, pagesize=A4,
                            topMargin=2 * cm, bottomMargin=2 * cm,
                            leftMargin=2 * cm, rightMargin=2 * cm)

    # Стили
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Title'],
        fontName=font_name,
        fontSize=16,
        alignment=1,  # Центр
        spaceAfter=20
    )

    heading_style = ParagraphStyle(
        'HeadingStyle',
        parent=styles['Heading2'],
        fontName=font_name,
        fontSize=12,
        alignment=1,
        spaceAfter=10
    )

    normal_style = ParagraphStyle(
        'NormalStyle',
        parent=styles['Normal'],
        fontName=font_name,
        fontSize=9,
        alignment=0  # Лево
    )

    # Собираем элементы документа
    story = []

    # Заголовок
    story.append(Paragraph(f"Результаты соревнования: {competition.title}", title_style))
    story.append(Spacer(1, 10))
    story.append(Paragraph(f"Дата: {competition.date}", normal_style))
    story.append(Paragraph(f"Место: {competition.location}", normal_style))
    story.append(Paragraph(f"Дата экспорта: {datetime.now().strftime('%d.%m.%Y %H:%M')}", normal_style))
    story.append(Spacer(1, 20))

    # Для каждого зачёта
    for entry in entries:
        # Заголовок зачёта
        story.append(Paragraph(f"{entry.get_discipline_display()} — {entry.get_sport_class_display()}", heading_style))

        # Получаем результаты
        results = Result.objects.filter(
            application_entry__entry=entry,
            application_entry__application__status=Application.STATUS_APPROVED,
            application_entry__application__payment_status=Application.PAYMENT_STATUS_PAID
        ).select_related(
            'application_entry__application__user',
            'application_entry__application__dog'
        ).order_by('place')

        if not results.exists():
            story.append(Paragraph("Нет участников в этом зачёте", normal_style))
            story.append(Spacer(1, 15))
            continue

        # Формируем данные для таблицы
        table_data = []

        # Заголовки таблицы в зависимости от дисциплины
        headers = ['Место', 'Спортсмен', 'Собака', 'Статус']
        if entry.discipline == 'bullseye':
            headers.extend(['Броски', 'Сумма', 'Кол-во'])
        elif entry.discipline == 'distance':
            headers.extend(['Попытка 1', 'Попытка 2', 'Попытка 3', 'Посл. шанс', 'Лучший'])
        elif entry.discipline == 'accuracy':
            headers.extend(['Броски', 'Сумма 5', 'Общая сумма'])

        table_data.append(headers)

        # Добавляем строки с результатами
        for result in results:
            app = result.application_entry.application
            row = [
                str(result.place or '—'),
                app.user.get_full_name() or app.user.username,
                app.dog.name,
                'Активен' if result.status == 'active' else 'Снят'
            ]

            if entry.discipline == 'bullseye':
                throws = result.data.get('throws', [])
                row.append(', '.join(map(str, throws)) if throws else '—')
                row.append(str(result.data.get('total', '—')))
                row.append(str(result.data.get('throw_count', '—')))
            elif entry.discipline == 'distance':
                attempts = result.data.get('attempts', [0, 0, 0])
                row.extend([str(attempts[0]), str(attempts[1]), str(attempts[2])])
                row.append(str(result.data.get('last_chance', 0)))
                row.append(str(result.data.get('best', 0)))
            elif entry.discipline == 'accuracy':
                throws = result.data.get('throws', [])
                row.append(', '.join(map(str, throws)) if throws else '—')
                row.append(str(result.data.get('top_five_sum', '—')))
                row.append(str(result.data.get('total_sum', '—')))

            table_data.append(row)

        # Создаём таблицу
        col_count = len(headers)
        available_width = doc.width
        col_widths = []

        for i, header in enumerate(headers):
            if i == 0:  # Место
                col_widths.append(40)
            elif i in (1, 2):  # Спортсмен, Собака
                col_widths.append(available_width * 0.25)
            elif i == 3:  # Статус
                col_widths.append(60)
            else:  # Остальные
                col_widths.append(available_width * 0.15)

        table = Table(table_data, colWidths=col_widths, repeatRows=1)

        # Стиль таблицы
        table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), font_name),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495e')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (1, 0), (2, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#bdc3c7')),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ]))

        story.append(table)
        story.append(Spacer(1, 20))

    # Собираем PDF
    doc.build(story)
    return response


def get_competition_entries_api(request, pk):
    """API для получения списка зачётов соревнования"""
    competition = get_object_or_404(Competition, pk=pk)
    entries_data = [{
        'id': entry.id,
        'discipline_name': entry.get_discipline_display(),
        'class_name': entry.get_sport_class_display(),
    } for entry in competition.entries.all()]
    return JsonResponse({'success': True, 'entries': entries_data})


# ============================================================
# ПУБЛИЧНЫЕ API
# ============================================================

def get_results_api(request, pk):
    """API получения результатов для организатора/судьи"""
    try:
        competition = get_object_or_404(Competition, pk=pk)
        discipline = request.GET.get('discipline')
        sport_class = request.GET.get('class')

        entry = competition.entries.filter(discipline=discipline, sport_class=sport_class).first()
        if not entry:
            return JsonResponse({'success': False, 'error': 'Зачёт не найден'})

        ensure_results_for_entry(entry)

        # Получаем все ApplicationEntry для этого зачёта
        application_entries = ApplicationEntry.objects.filter(entry=entry).select_related(
            'application__user', 'application__dog'
        )

        results_data = []

        for app_entry in application_entries:
            application = app_entry.application

            # Проверяем, что заявка одобрена и оплачена
            if application.status != Application.STATUS_APPROVED or application.payment_status != Application.PAYMENT_STATUS_PAID:
                continue

            # Получаем или создаём Result
            result, created = Result.objects.get_or_create(
                application_entry=app_entry,
                defaults={'data': {}, 'sort_value': 0, 'place': None, 'status': Result.STATUS_ACTIVE}
            )

            results_data.append({
                'id': result.id,
                'start_order': result.start_order,
                'place': result.place if not app_entry.is_out_of_class else None,
                'is_out_of_class': app_entry.is_out_of_class,  # ← КЛЮЧЕВОЕ ПОЛЕ
                'user_name': application.user.get_full_name() or application.user.username,
                'dog_name': application.dog.name,
                'data': result.data,
                'sort_value': result.sort_value,
                'status': result.status
            })

        # Сортируем: сначала основная таблица (по sort_value), потом вне зачёта
        results_data.sort(key=lambda x: (x['is_out_of_class'], -x['sort_value']))

        return JsonResponse({'success': True, 'results': results_data})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


def get_results_api_public(request, pk):
    """Публичное API для получения результатов"""
    return get_results_api(request, pk)


def get_competitions_all_api(request):
    """API для получения всех опубликованных соревнований"""
    competitions_data = []
    for comp in Competition.objects.filter(status=Competition.STATUS_PUBLISHED).order_by('-date'):
        disciplines = list(comp.entries.values_list('discipline', flat=True).distinct())
        classes = list(comp.entries.values_list('sport_class', flat=True).distinct())
        competitions_data.append({
            'id': comp.id,
            'title': comp.title,
            'date': comp.date.isoformat(),
            'location': comp.location,
            'city': comp.city or '',
            'registration_deadline': comp.registration_deadline.isoformat(),
            'location_type': comp.location_type,
            'is_registration_open': comp.is_registration_open(),
            'organizer_id': comp.organizer.id,
            'has_disciplines': len(disciplines) > 0,
            'disciplines': disciplines,
            'has_classes': len(classes) > 0,
            'classes': classes,
        })
    return JsonResponse({'success': True, 'competitions': competitions_data})


def get_filter_options_api(request):
    """API для получения доступных городов и организаторов"""
    competitions = Competition.objects.filter(status=Competition.STATUS_PUBLISHED)
    cities = sorted(competitions.exclude(city='').exclude(city__isnull=True).values_list('city', flat=True).distinct())

    organizers = []
    for org_id in competitions.values_list('organizer', flat=True).distinct():
        try:
            user = User.objects.get(id=org_id)
            organizers.append({'id': user.id, 'name': user.get_full_name() or user.username})
        except User.DoesNotExist:
            pass

    return JsonResponse({'success': True, 'cities': cities, 'organizers': organizers})


def api_get_judges(request):
    """API для получения списка судей (организаторов)"""
    if not request.user.is_authenticated:
        return JsonResponse({'success': False, 'error': 'Не авторизован'})

    judges = [{
        'id': org.id,
        'name': org.get_full_name() or org.username,
        'is_current_user': org.id == request.user.id
    } for org in User.objects.filter(profile__is_organizer=True)]

    return JsonResponse({'success': True, 'judges': judges})


# ============================================================
# РАБОТА С ФАЙЛАМИ
# ============================================================

@csrf_exempt
def upload_competition_file(request):
    """Загрузка файла для соревнования (временное хранение)"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Метод не поддерживается'})

    if not request.user.is_authenticated or not request.user.profile.is_organizer:
        return JsonResponse({'success': False, 'error': 'Недостаточно прав'})

    file = request.FILES.get('file')
    if not file:
        return JsonResponse({'success': False, 'error': 'Файл не передан'})

    if file.size > 10 * 1024 * 1024:
        return JsonResponse({'success': False, 'error': 'Файл слишком большой (максимум 10MB)'})

    try:
        ext = file.name.split('.')[-1]
        original_name = file.name
        # Сохраняем оригинальное имя в пути через ___
        unique_name = f"temp/{uuid.uuid4().hex}___{original_name}"
        path = default_storage.save(unique_name, ContentFile(file.read()))

        return JsonResponse({
            'success': True,
            'file': {
                'name': original_name,
                'path': path,
                'size': file.size,
                'type': file.content_type
            }
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Ошибка сохранения: {str(e)}'})


# ============================================================
# ВСПОМОГАТЕЛЬНЫЕ
# ============================================================

def get_dog_card(request):
    """Получение карточки собаки для модального окна"""
    dog_id = request.GET.get('dog_id')
    show_select_button = request.GET.get('show_select_button') == 'true'
    show_select_other_button = request.GET.get('show_select_other_button') == 'true'

    if not dog_id:
        return JsonResponse({'html': '', 'success': False, 'error': 'ID собаки не указан'})

    try:
        dog = Dog.objects.get(id=dog_id)
        html = render_to_string('users/dog_card_for_select.html', {
            'dog': dog,
            'show_select_button': show_select_button,
            'show_select_other_button': show_select_other_button,
        })
        return JsonResponse({'html': html, 'success': True})
    except Dog.DoesNotExist:
        return JsonResponse({'html': '', 'success': False, 'error': 'Собака не найдена'})


def filter_competitions(request):
    """API для фильтрации соревнований и рендера карточек"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Метод не поддерживается'}, status=405)

    try:
        data = json.loads(request.body)

        # Базовый запрос
        competitions = Competition.objects.filter(status=Competition.STATUS_PUBLISHED)

        # Фильтр по поиску
        search = data.get('search', '').strip()
        if search:
            competitions = competitions.filter(
                Q(title__iregex=search) |
                Q(title__icontains=search)
            )

        # Фильтр по дисциплинам
        disciplines = data.get('disciplines', [])
        if disciplines:
            competitions = competitions.filter(
                entries__discipline__in=disciplines
            ).distinct()

        # Фильтр по классам
        classes = data.get('classes', [])
        if classes:
            competitions = competitions.filter(
                entries__sport_class__in=classes
            ).distinct()

        # Фильтр по периоду
        period = data.get('period')
        from django.utils import timezone
        today = timezone.now().date()
        if period == 'upcoming':
            competitions = competitions.filter(date__gte=today)
        elif period == 'past':
            competitions = competitions.filter(date__lt=today)

        # Фильтр по дате (диапазон)
        date_from = data.get('date_from')
        if date_from:
            competitions = competitions.filter(date__gte=date_from)
        date_to = data.get('date_to')
        if date_to:
            competitions = competitions.filter(date__lte=date_to)

        # Фильтр по типу площадки
        venues = data.get('venues', [])
        if venues:
            competitions = competitions.filter(location_type__in=venues)

        # Фильтр по городу
        cities = data.get('cities', [])
        if cities:
            competitions = competitions.filter(city__in=cities)

        # Фильтр по организатору
        organizers = data.get('organizers', [])
        if organizers:
            competitions = competitions.filter(organizer_id__in=organizers)

        # Рендерим карточки
        html = ''
        for comp in competitions.order_by('-date'):
            html += render_to_string('competitions/competition_card.html', {'comp': comp})

        return JsonResponse({'success': True, 'html': html})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# competitions/views.py
# competitions/views.py

def get_application_form_html(request):
    """Возвращает HTML формы заявки с поддержкой 'Вне зачёта'"""
    from django.template.loader import render_to_string
    from users.models import Dog
    from .models import Competition

    dog_id = request.GET.get('dog_id')
    form_id = request.GET.get('form_id')
    competition_id = request.GET.get('competition_id')

    try:
        dog = Dog.objects.get(id=dog_id)
        competition = Competition.objects.get(
            id=competition_id,
            status=Competition.STATUS_PUBLISHED
        )

        entries = competition.entries.all().select_related('competition')

        discipline_groups = []
        disciplines = entries.values_list('discipline', flat=True).distinct()

        class_order = ['novice', 'progress', 'open']

        for discipline_code in disciplines:
            discipline_entries = entries.filter(discipline=discipline_code)
            discipline_name = discipline_entries.first().get_discipline_display()

            class_groups = []
            classes = discipline_entries.values_list('sport_class', flat=True).distinct()

            for class_code in classes:
                class_entries = discipline_entries.filter(sport_class=class_code)
                class_name = class_entries.first().get_sport_class_display()
                entry = class_entries.first()

                judge_name = None
                first_judge = entry.assigned_judges.first()
                if first_judge:
                    judge_name = first_judge.user.get_full_name() or first_judge.user.username
                elif entry.judge_name:
                    judge_name = entry.judge_name

                class_groups.append({
                    'class_code': class_code,
                    'class_name': class_name,
                    'entry': {
                        'id': entry.id,
                        'judge_name': judge_name or 'Не указан',
                        'fee': str(entry.fee),
                        'can_enter_without_class': entry.can_enter_without_class,
                    }
                })

            # Сортируем: новички → прогресс → открытый
            class_groups.sort(key=lambda x: class_order.index(x['class_code']) if x['class_code'] in class_order else 99)

            discipline_groups.append({
                'discipline_name': discipline_name,
                'classes': class_groups
            })

        entries_html = render_to_string('competitions/entries_checkboxes.html', {
            'discipline_groups': discipline_groups
        })

        html = render_to_string('competitions/application_form.html', {
            'dog': dog,
            'user': request.user,
            'form_id': form_id,
            'entries_html': entries_html,
        })

        return JsonResponse({'success': True, 'html': html})

    except Dog.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Собака не найдена'})
    except Competition.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Соревнование не найдено'})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': f'Ошибка сервера: {str(e)}'})


def get_competition_card_html(request):
    """Возвращает HTML карточки соревнования"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Метод не поддерживается'})

    try:
        data = json.loads(request.body)
        comp_data = data.get('competition')

        # Получаем объект Competition из базы
        competition = Competition.objects.get(id=comp_data['id'])

        # Добавляем недостающие поля
        comp_data['date_display'] = competition.date.strftime('%d.%m.%Y') if competition.date else None
        comp_data['registration_deadline_display'] = competition.registration_deadline.strftime(
            '%d.%m.%Y') if competition.registration_deadline else None
        comp_data['is_finished'] = competition.is_finished()
        comp_data['is_ongoing'] = competition.is_ongoing()
        comp_data['is_registration_open'] = competition.is_registration_open()
        comp_data['get_unique_disciplines_display'] = competition.get_unique_disciplines_display()

        # КЛЮЧЕВОЕ: преобразуем date в объект даты, чтобы фильтр date работал
        from datetime import datetime
        if comp_data.get('date'):
            comp_data['date'] = datetime.fromisoformat(comp_data['date']).date()
        if comp_data.get('registration_deadline'):
            comp_data['registration_deadline'] = datetime.fromisoformat(comp_data['registration_deadline']).date()

        html = render_to_string('competitions/competition_card.html', {'comp': comp_data})
        return JsonResponse({'success': True, 'html': html})
    except Competition.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Соревнование не найдено'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def update_start_order(request, pk):
    """Сохранение стартового порядка"""
    competition = get_object_or_404(Competition, pk=pk)

    if not (competition.organizer == request.user or
            Entry.objects.filter(competition=competition, assigned_judges__user=request.user).exists()):
        return JsonResponse({'success': False, 'error': 'Нет прав'}, status=403)

    if request.method == 'POST':
        data = json.loads(request.body)
        order_list = data.get('order', [])
        for item in order_list:
            Result.objects.filter(id=item['result_id']).update(start_order=item['start_order'])
        return JsonResponse({'success': True})

    return JsonResponse({'success': False, 'error': 'Метод не поддерживается'})